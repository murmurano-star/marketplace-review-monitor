from __future__ import annotations

import asyncio
import html
import json
import re
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote, unquote, urlparse, parse_qs

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_XLSX = Path('reports/chinese_vin_open_sources.xlsx')
OUT_JSON = Path('reports/chinese_vin_open_sources.json')
VIN_RE = re.compile(r'(?<![A-Z0-9])[A-HJ-NPR-Z0-9]{17}(?![A-Z0-9])', re.I)

BRANDS = {
    'Chery / EXEED / OMODA / JAECOO / Jetour': ['LVV', 'LVM'],
    'Geely': ['L6T', 'LB3'],
    'GWM / Haval / Tank / Ora / Wey': ['LGW'],
    'Changan': ['LS5'],
    'BYD': ['LC0', 'LPE'],
    'GAC / Trumpchi': ['LMG'],
    'SAIC MG / Roewe': ['LSJ'],
    'SAIC Maxus': ['LSF', 'LSH', 'LSK'],
    'Wuling / Baojun': ['LZW'],
    'DFSK / Dongfeng Sokon': ['LVZ'],
    'BAIC': ['LNB'],
    'Foton': ['LVA', 'LVB', 'LVC'],
    'JAC': ['LJ1'],
    'FAW / Bestune / Hongqi': ['LFB', 'LFP'],
    'Lifan': ['LLV'],
    'Qoros': ['LLN'],
    'SWM': ['LM6'],
    'Landwind': ['LVX'],
    'ZX Auto': ['LTA'],
}
PREFIX_TO_BRAND = {p: b for b, ps in BRANDS.items() for p in ps}
BRAND_TERMS = {
    'Chery / EXEED / OMODA / JAECOO / Jetour': ['Chery','EXEED','OMODA','JAECOO','Jetour'],
    'Geely': ['Geely'],
    'GWM / Haval / Tank / Ora / Wey': ['Haval','Great Wall','GWM','Tank','Ora','Wey'],
    'Changan': ['Changan'],
    'BYD': ['BYD'],
    'GAC / Trumpchi': ['GAC','Trumpchi'],
    'SAIC MG / Roewe': ['MG','Roewe'],
    'SAIC Maxus': ['Maxus'],
    'Wuling / Baojun': ['Wuling','Baojun'],
    'DFSK / Dongfeng Sokon': ['DFSK','Dongfeng Sokon'],
    'BAIC': ['BAIC'],
    'Foton': ['Foton'],
    'JAC': ['JAC'],
    'FAW / Bestune / Hongqi': ['FAW','Bestune','Hongqi'],
    'Lifan': ['Lifan'],
    'Qoros': ['Qoros'],
    'SWM': ['SWM'],
    'Landwind': ['Landwind'],
    'ZX Auto': ['ZX Auto'],
}

TRANSLIT = {**{str(i): i for i in range(10)}, **dict(zip('ABCDEFGHJKLMNPRSTUVWXYZ', [1,2,3,4,5,6,7,8,1,2,3,4,5,7,9,2,3,4,5,6,7,8,9]))}
WEIGHTS = [8,7,6,5,4,3,2,10,0,9,8,7,6,5,4,3,2]

def valid_check_digit(vin: str) -> bool:
    try:
        total = sum(TRANSLIT[c] * w for c, w in zip(vin, WEIGHTS))
    except KeyError:
        return False
    expected = 'X' if total % 11 == 10 else str(total % 11)
    return vin[8] == expected


def brand_for(vin: str) -> str | None:
    for n in (3,):
        if vin[:n] in PREFIX_TO_BRAND:
            return PREFIX_TO_BRAND[vin[:n]]
    return None


def clean_url(url: str) -> str:
    url = html.unescape(url.strip())
    if 'uddg=' in url:
        try:
            url = unquote(parse_qs(urlparse(url).query)['uddg'][0])
        except Exception:
            pass
    return url.rstrip(').,]')


def extract_urls(text: str) -> list[str]:
    urls = []
    for u in re.findall(r'https?://[^\s<>"\']+', text):
        u = clean_url(u)
        if any(x in u for x in ['duckduckgo.com/y.js','google.com/search','bing.com/search']):
            continue
        urls.append(u)
    for u in re.findall(r'\]\((https?://[^)]+)\)', text):
        urls.append(clean_url(u))
    out=[]; seen=set()
    for u in urls:
        if u not in seen:
            seen.add(u); out.append(u)
    return out


def contexts(text: str, vin: str) -> str:
    idx = text.find(vin)
    if idx < 0:
        return ''
    return re.sub(r'\s+', ' ', text[max(0, idx-180):idx+len(vin)+220]).strip()


async def fetch(client: httpx.AsyncClient, url: str, timeout=25) -> tuple[int, str]:
    try:
        r = await client.get(url, timeout=timeout, follow_redirects=True)
        return r.status_code, r.text
    except Exception as exc:
        return 0, f'ERROR {type(exc).__name__}: {exc}'


async def search_one(client, query: str) -> tuple[str, list[str], str]:
    providers = [
        f'https://r.jina.ai/http://www.google.com/search?q={quote(query)}',
        f'https://r.jina.ai/http://www.bing.com/search?q={quote(query)}',
        f'https://html.duckduckgo.com/html/?q={quote(query)}',
    ]
    combined=[]; urls=[]
    for u in providers:
        status, text = await fetch(client, u)
        combined.append(f'\nSOURCE {u} STATUS {status}\n{text[:200000]}')
        urls.extend(extract_urls(text))
        await asyncio.sleep(0.2)
    blob='\n'.join(combined)
    return query, urls[:30], blob


async def read_page(client, url: str) -> tuple[str,int,str]:
    candidates=[url]
    if not url.startswith('https://r.jina.ai/'):
        candidates.insert(0, 'https://r.jina.ai/http://' + url.split('://',1)[-1])
    for u in candidates:
        status,text=await fetch(client,u,30)
        if status==200 and len(text)>100:
            return url,status,text
    return url,status,text


async def main():
    queries=[]
    for brand, terms in BRAND_TERMS.items():
        prefixes=BRANDS[brand]
        for term in terms[:2]:
            queries += [
                f'"{term}" VIN',
                f'"{term}" "VIN-код"',
                f'site:drive2.ru "{term}" VIN',
                f'site:sudact.ru "{term}" VIN',
            ]
        for p in prefixes:
            queries += [f'"{p}" VIN automobile', f'"{p}" "{terms[0]}"']
    # High-yield public source categories.
    queries += [
        'Росстандарт отзыв Chery VIN приложение', 'Росстандарт отзыв Haval VIN приложение',
        'Росстандарт отзыв Geely VIN приложение', 'Росстандарт отзыв Changan VIN приложение',
        'auction Chery full VIN', 'auction Haval full VIN', 'auction Geely full VIN',
        'auction Changan full VIN', 'court Chery VIN', 'court Haval VIN',
    ]
    queries=list(dict.fromkeys(queries))
    headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36','Accept-Language':'ru-RU,ru;q=0.9,en;q=0.7'}
    limits=httpx.Limits(max_connections=15,max_keepalive_connections=10)
    async with httpx.AsyncClient(headers=headers,limits=limits,http2=True) as client:
        sem=asyncio.Semaphore(5)
        async def s(q):
            async with sem:
                return await search_one(client,q)
        search_results=await asyncio.gather(*(s(q) for q in queries))
        candidate_urls=[]
        search_blobs=[]
        records=[]
        for query,urls,blob in search_results:
            search_blobs.append((query,blob))
            candidate_urls.extend(urls)
            for m in VIN_RE.finditer(blob.upper()):
                vin=m.group(0).upper(); brand=brand_for(vin)
                if brand:
                    records.append({'vin':vin,'brand':brand,'source':'search result','url':'','query':query,'context':contexts(blob.upper(),vin),'check_digit_valid':valid_check_digit(vin)})
        # Prefer domains likely to quote vehicle identifiers; cap for run time.
        priority=['rst.gov.ru','rosstandart','sudact.ru','drive2.ru','auto.ru','avito.ru','drom.ru','auction','bid','vin','car','forum','pdf','gov.ru']
        def score(u): return sum(3 for x in priority if x in u.lower()) - len(u)/10000
        unique_urls=sorted(set(candidate_urls),key=score,reverse=True)[:300]
        async def rp(u):
            async with sem:
                return await read_page(client,u)
        pages=await asyncio.gather(*(rp(u) for u in unique_urls))
        for url,status,text in pages:
            upper=text.upper()
            for m in VIN_RE.finditer(upper):
                vin=m.group(0).upper(); brand=brand_for(vin)
                if not brand: continue
                records.append({'vin':vin,'brand':brand,'source':'web page','url':url,'query':'','context':contexts(upper,vin),'check_digit_valid':valid_check_digit(vin),'http_status':status})
    # Deduplicate source mentions, retain exact citations.
    dedup={}
    for r in records:
        key=(r['vin'],r.get('url',''),r.get('query',''))
        dedup[key]=r
    records=list(dedup.values())
    # Require a mapped Chinese WMI; preserve invalid check digit with warning because public pages can contain typos.
    records.sort(key=lambda r:(r['brand'],r['vin'],r.get('url','')))
    OUT_XLSX.parent.mkdir(parents=True,exist_ok=True)
    OUT_JSON.write_text(json.dumps(records,ensure_ascii=False,indent=2),encoding='utf-8')
    wb=Workbook(); ws=wb.active; ws.title='Упоминания VIN'
    headers=['VIN','Бренд / производитель','Контрольная цифра VIN','Тип источника','URL источника','Поисковый запрос','Контекст цитирования']
    ws.append(headers)
    for r in records:
        ws.append([r['vin'],r['brand'],'корректна' if r['check_digit_valid'] else 'не прошла проверку',r['source'],r.get('url',''),r.get('query',''),r.get('context','')])
    sm=wb.create_sheet('Сводка')
    sm.append(['Бренд / производитель','Уникальных VIN','Упоминаний'])
    by_brand=defaultdict(list)
    for r in records: by_brand[r['brand']].append(r)
    for b,rs in sorted(by_brand.items()): sm.append([b,len({r['vin'] for r in rs}),len(rs)])
    sm.append([]); sm.append(['Всего уникальных VIN',len({r['vin'] for r in records})]); sm.append(['Всего цитируемых упоминаний',len(records)]); sm.append(['Проверенных URL',len(unique_urls)])
    fill=PatternFill('solid',fgColor='F58220')
    for sh in wb.worksheets:
        for c in sh[1]: c.fill=fill; c.font=Font(bold=True,color='FFFFFF'); c.alignment=Alignment(wrap_text=True,vertical='top')
        sh.freeze_panes='A2'; sh.auto_filter.ref=sh.dimensions
        for row in sh.iter_rows():
            for c in row: c.alignment=Alignment(wrap_text=True,vertical='top')
        for col in range(1,sh.max_column+1):
            maxlen=max((len(str(sh.cell(r,col).value or '')) for r in range(1,min(sh.max_row,250)+1)),default=10)
            sh.column_dimensions[get_column_letter(col)].width=min(max(maxlen+2,12),65)
    ws.column_dimensions['E'].width=55; ws.column_dimensions['G'].width=90
    wb.save(OUT_XLSX)
    print(json.dumps({'records':len(records),'unique_vins':len({r['vin'] for r in records}),'urls':len(unique_urls),'xlsx':str(OUT_XLSX)},ensure_ascii=False))

if __name__=='__main__': asyncio.run(main())
