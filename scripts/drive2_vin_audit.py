from __future__ import annotations

import argparse
import asyncio
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import quote, urljoin, urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import BrowserContext, Page, async_playwright


VIN_RE = re.compile(r"(?<![A-HJ-NPR-Z0-9])([A-HJ-NPR-Z0-9](?:[\s-]?[A-HJ-NPR-Z0-9]){16})(?![A-HJ-NPR-Z0-9])", re.I)
FRAME_RE = re.compile(r"(?<![A-Z0-9])([A-Z]{1,6}[A-Z0-9]{1,10}-[A-Z0-9]{4,12})(?![A-Z0-9])", re.I)
PAGE_RE = re.compile(r"[?&]page=(\d+)")

BLOCK_MARKERS = (
    "captcha", "капча", "я не робот", "подтвердите, что вы не робот",
    "доступ ограничен", "access denied", "temporarily blocked", "слишком много запросов",
)

RAVENOL_NOT_FOUND = (
    "ничего не найдено", "не найден", "нет результатов", "not found",
    "не удалось определить", "автомобиль не определен", "автомобиль не определён",
)

MANUFACTURER_ALIASES = {
    "mercedes-benz": "Mercedes-Benz", "mercedes": "Mercedes-Benz",
    "volkswagen": "Volkswagen", "vw": "Volkswagen",
    "hyundai": "Hyundai", "kia": "Kia", "toyota": "Toyota", "lexus": "Lexus",
    "nissan": "Nissan", "infiniti": "Infiniti", "renault": "Renault", "dacia": "Dacia",
    "lada": "LADA", "vaz": "LADA", "ваз": "LADA", "uaz": "УАЗ", "gaz": "ГАЗ",
    "bmw": "BMW", "mini": "MINI", "audi": "Audi", "skoda": "Škoda", "seat": "SEAT",
    "ford": "Ford", "chevrolet": "Chevrolet", "opel": "Opel", "vauxhall": "Vauxhall",
    "peugeot": "Peugeot", "citroen": "Citroën", "ds": "DS Automobiles",
    "mitsubishi": "Mitsubishi", "mazda": "Mazda", "subaru": "Subaru", "honda": "Honda",
    "acura": "Acura", "suzuki": "Suzuki", "isuzu": "Isuzu", "daihatsu": "Daihatsu",
    "geely": "Geely", "chery": "Chery", "haval": "Haval", "great-wall": "Great Wall",
    "gac": "GAC", "jac": "JAC", "exeed": "EXEED", "omoda": "OMODA", "jetour": "JETOUR",
    "moskvich": "Москвич", "moskvich-3": "Москвич", "byd": "BYD", "zeekr": "ZEEKR",
    "volvo": "Volvo", "saab": "Saab", "fiat": "Fiat", "alfa-romeo": "Alfa Romeo",
    "porsche": "Porsche", "land-rover": "Land Rover", "jaguar": "Jaguar", "jeep": "Jeep",
    "chrysler": "Chrysler", "dodge": "Dodge", "cadillac": "Cadillac", "buick": "Buick",
    "tesla": "Tesla", "smart": "smart", "ssangyong": "SsangYong", "daewoo": "Daewoo",
}

# Conservative WMI fallback. It is only used when Ravenol was queried but did not expose a readable brand.
WMI_PREFIXES = {
    "WVW": "Volkswagen", "WVG": "Volkswagen", "WUA": "Audi", "WAU": "Audi",
    "WBA": "BMW", "WBS": "BMW", "WBY": "BMW", "WDD": "Mercedes-Benz", "WDB": "Mercedes-Benz",
    "VF1": "Renault", "VF3": "Peugeot", "VF7": "Citroën", "VR1": "DS Automobiles",
    "TMB": "Škoda", "TM9": "Škoda", "VSS": "SEAT", "ZFA": "Fiat", "ZAR": "Alfa Romeo",
    "SAL": "Land Rover", "SAJ": "Jaguar", "WP0": "Porsche", "YV1": "Volvo",
    "JTD": "Toyota", "JTE": "Toyota", "JT1": "Toyota", "JTH": "Lexus",
    "JN1": "Nissan", "JNK": "Infiniti", "JMZ": "Mazda", "JMB": "Mitsubishi",
    "JHM": "Honda", "JH4": "Acura", "JS2": "Suzuki", "JF1": "Subaru", "JF2": "Subaru",
    "KMH": "Hyundai", "KNA": "Kia", "KNE": "Kia", "KL1": "Chevrolet", "KPT": "SsangYong",
    "Z94": "Hyundai", "XWE": "Hyundai", "XW8": "Volkswagen", "XTA": "LADA",
    "XTT": "УАЗ", "X96": "ГАЗ", "XUF": "Chevrolet", "XUU": "Chevrolet", "X7L": "Renault",
    "1HG": "Honda", "1N4": "Nissan", "1NX": "Toyota", "1FA": "Ford", "1FT": "Ford",
    "1G1": "Chevrolet", "1GC": "Chevrolet", "1C4": "Jeep", "1C3": "Chrysler",
    "2C3": "Chrysler", "2HG": "Honda", "2T1": "Toyota", "3VW": "Volkswagen",
    "5YJ": "Tesla", "5UX": "BMW", "5N1": "Nissan", "5NP": "Hyundai",
    "LVS": "Ford", "LVV": "Chery", "LGW": "Great Wall", "LHG": "Honda",
    "LDC": "Dongfeng", "LSG": "Chevrolet", "LBV": "BMW", "LVR": "Changan",
}


@dataclass
class Mention:
    code: str
    code_type: str
    page: int
    comment_id: str
    author: str
    text: str
    source_url: str


@dataclass
class RavenolResult:
    code: str
    manufacturer: str = ""
    model: str = ""
    engine: str = ""
    year: str = ""
    status: str = ""
    source_method: str = ""
    result_url: str = ""
    response_excerpt: str = ""


def clean_code(raw: str) -> str:
    return re.sub(r"[\s-]", "", raw).upper()


def extract_codes(text: str) -> list[tuple[str, str]]:
    normalized = text.upper().replace("\u200b", " ").replace("\xa0", " ")
    found: list[tuple[str, str]] = []
    seen: set[str] = set()
    for match in VIN_RE.finditer(normalized):
        code = clean_code(match.group(1))
        if len(code) == 17 and code not in seen:
            seen.add(code)
            found.append((code, "VIN"))
    # Frame/chassis identifiers are accepted only when the surrounding comment explicitly refers to VIN/frame/body.
    if re.search(r"\b(vin|frame|вин|фрейм|кузов|шасси)\b", normalized, re.I):
        for match in FRAME_RE.finditer(normalized):
            code = match.group(1).upper()
            if code not in seen:
                seen.add(code)
                found.append((code, "Frame"))
    return found


def blocked(text: str) -> bool:
    low = text.lower()
    return any(marker in low for marker in BLOCK_MARKERS)


async def new_page(context: BrowserContext) -> Page:
    page = await context.new_page()
    await page.route("**/*", lambda route: route.abort() if route.request.resource_type in {"image", "media", "font"} else route.continue_())
    return page


async def load_drive2_page(page: Page, url: str, timeout_ms: int = 45000) -> tuple[str, str]:
    response = await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
    status = response.status if response else 0
    html = await page.content()
    body = await page.locator("body").inner_text(timeout=10000)
    if status >= 400:
        raise RuntimeError(f"Drive2 returned HTTP {status}: {url}")
    if blocked((await page.title()) + "\n" + body):
        raise RuntimeError(f"Drive2 anti-bot/CAPTCHA page detected: {url}")
    return html, body


async def collect_drive2_mentions(context: BrowserContext, base_url: str, requested_last_page: int) -> tuple[list[Mention], list[dict]]:
    page = await new_page(context)
    logs: list[dict] = []
    mentions: list[Mention] = []
    try:
        html, _ = await load_drive2_page(page, base_url)
        page_numbers = [int(x) for x in PAGE_RE.findall(html)]
        detected_last = max(page_numbers, default=1)
        last_page = max(requested_last_page, detected_last)
        print(f"Detected pagination: {detected_last}; processing pages 1..{last_page}", flush=True)

        empty_streak = 0
        for number in range(1, last_page + 1):
            url = base_url if number == 1 else f"{base_url}?page={number}#comments"
            try:
                _, body = await load_drive2_page(page, url)
                comments = page.locator("div.c-comment[data-role='comment'], article.c-comment[data-role='comment']")
                count = await comments.count()
                # Fallback for markup variants.
                if count == 0:
                    comments = page.locator("[data-role='comment']")
                    count = await comments.count()
                logs.append({"page": number, "url": url, "status": "ok", "comments": count, "note": ""})
                print(f"Drive2 page {number}/{last_page}: {count} comment nodes", flush=True)
                if count == 0:
                    empty_streak += 1
                else:
                    empty_streak = 0
                for idx in range(count):
                    node = comments.nth(idx)
                    comment_id = (await node.get_attribute("id")) or ""
                    comment_id = comment_id.lstrip("a")
                    text_loc = node.locator("[data-slot='comment.text'], .c-comment__text").first
                    try:
                        text = (await text_loc.inner_text(timeout=3000)).strip()
                    except Exception:
                        text = (await node.inner_text()).strip()
                    author = ""
                    for selector in (".c-username [itemprop='name']", ".c-username", "[itemprop='name']"):
                        loc = node.locator(selector).first
                        if await loc.count():
                            try:
                                author = (await loc.inner_text(timeout=1000)).strip()
                                if author:
                                    break
                            except Exception:
                                pass
                    for code, code_type in extract_codes(text):
                        mentions.append(Mention(code, code_type, number, comment_id, author, text, url))
                # Do not silently truncate a known 144-page thread, but stop if the site starts returning empty pages beyond it.
                if number > requested_last_page and empty_streak >= 3:
                    print("Stopping after 3 empty pages beyond requested range", flush=True)
                    break
            except Exception as exc:
                logs.append({"page": number, "url": url, "status": "error", "comments": 0, "note": str(exc)})
                print(f"ERROR Drive2 page {number}: {exc}", file=sys.stderr, flush=True)
                if "CAPTCHA" in str(exc) or "anti-bot" in str(exc):
                    raise
        return mentions, logs
    finally:
        await page.close()


def normalize_manufacturer(value: str) -> str:
    value = re.sub(r"\s+", " ", value).strip(" \t\r\n:;|–—-")
    low = value.lower()
    if low in MANUFACTURER_ALIASES:
        return MANUFACTURER_ALIASES[low]
    for key, canonical in MANUFACTURER_ALIASES.items():
        if re.search(rf"(?<![a-z]){re.escape(key)}(?![a-z])", low):
            return canonical
    if 1 < len(value) <= 40:
        return value
    return ""


def parse_label(body: str, labels: Iterable[str]) -> str:
    for label in labels:
        patterns = [
            rf"{label}\s*[:：]\s*([^\n\r|]{{1,100}})",
            rf"{label}\s*\n\s*([^\n\r|]{{1,100}})",
        ]
        for pattern in patterns:
            match = re.search(pattern, body, re.I)
            if match:
                return re.sub(r"\s+", " ", match.group(1)).strip()
    return ""


def wmi_fallback(code: str) -> str:
    if len(code) != 17:
        return ""
    return WMI_PREFIXES.get(code[:3], "")


async def decode_with_ravenol(context: BrowserContext, code: str, debug_dir: Path) -> RavenolResult:
    page = await new_page(context)
    search_url = f"https://podbor.ravenol.ru/search/?q={quote(code)}"
    result = RavenolResult(code=code, result_url=search_url)
    try:
        response = await page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
        status_code = response.status if response else 0
        body = (await page.locator("body").inner_text(timeout=15000)).strip()
        final_url = page.url
        result.result_url = final_url
        result.response_excerpt = re.sub(r"\s+", " ", body)[:600]

        if status_code >= 400:
            result.status = f"HTTP {status_code}"
            return result
        if blocked((await page.title()) + "\n" + body):
            result.status = "Ravenol CAPTCHA/anti-bot"
            return result

        result.manufacturer = normalize_manufacturer(parse_label(body, ("Марка", "Производитель", "Бренд")))
        result.model = parse_label(body, ("Модель", "Model"))
        result.engine = parse_label(body, ("Двигатель", "Код двигателя", "Engine"))
        result.year = parse_label(body, ("Год выпуска", "Модельный год", "Год", "Year"))

        # Ravenol result URLs normally include /1-cars/<id>-<manufacturer>/...
        path = urlparse(final_url).path.lower()
        slug_match = re.search(r"/1-cars/\d+-([^/]+)/", path)
        if not result.manufacturer and slug_match:
            slug = slug_match.group(1).strip("-")
            result.manufacturer = normalize_manufacturer(slug.replace("-", " "))
            result.source_method = "Ravenol result URL"

        # Inspect breadcrumb and model links when labels are not present.
        if not result.manufacturer:
            candidates = []
            for selector in ("nav a", ".breadcrumb a", "a[href*='/1-cars/']", "h1", "h2"):
                loc = page.locator(selector)
                for i in range(min(await loc.count(), 30)):
                    try:
                        text = re.sub(r"\s+", " ", await loc.nth(i).inner_text(timeout=500)).strip()
                    except Exception:
                        continue
                    if text:
                        candidates.append(text)
            for candidate in candidates:
                brand = normalize_manufacturer(candidate)
                if brand in set(MANUFACTURER_ALIASES.values()):
                    result.manufacturer = brand
                    result.source_method = "Ravenol page element"
                    break

        low_body = body.lower()
        if result.manufacturer:
            result.status = "Распознан Ravenol"
            if not result.source_method:
                result.source_method = "Ravenol page fields"
        elif any(marker in low_body for marker in RAVENOL_NOT_FOUND):
            result.status = "Не найден Ravenol"
        else:
            # Save the response to make ambiguous cases auditable.
            debug_dir.mkdir(parents=True, exist_ok=True)
            (debug_dir / f"ravenol_{code}.txt").write_text(f"URL: {final_url}\n\n{body}", encoding="utf-8")
            result.status = "Ответ Ravenol неоднозначен"

        if not result.manufacturer:
            fallback = wmi_fallback(code)
            if fallback:
                result.manufacturer = fallback
                result.source_method = "WMI fallback after Ravenol query"
                result.status += "; производитель по WMI"
        return result
    except Exception as exc:
        result.status = f"Ошибка Ravenol: {exc}"
        fallback = wmi_fallback(code)
        if fallback:
            result.manufacturer = fallback
            result.source_method = "WMI fallback after Ravenol error"
        return result
    finally:
        await page.close()


async def decode_all(context: BrowserContext, codes: list[str], debug_dir: Path) -> dict[str, RavenolResult]:
    results: dict[str, RavenolResult] = {}
    # Sequential requests are intentional: they reduce load on the public catalog and make blocking less likely.
    for index, code in enumerate(codes, 1):
        print(f"Ravenol {index}/{len(codes)}: {code}", flush=True)
        results[code] = await decode_with_ravenol(context, code, debug_dir)
        await asyncio.sleep(0.35)
    return results


def style_sheet(ws, freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        values = [str(c.value) if c.value is not None else "" for c in column_cells[:200]]
        max_len = max((len(v) for v in values), default=10)
        width = min(max(max_len + 2, 11), 55)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_excel(output: Path, base_url: str, mentions: list[Mention], logs: list[dict], decoded: dict[str, RavenolResult]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "VIN по производителям"

    by_code: dict[str, list[Mention]] = defaultdict(list)
    for mention in mentions:
        by_code[mention.code].append(mention)

    headers = [
        "VIN / Frame", "Тип", "Автопроизводитель", "Модель", "Двигатель", "Год",
        "Статус проверки", "Источник определения", "URL результата Ravenol",
        "Количество упоминаний", "Страницы Drive2", "Первый автор", "Первый комментарий",
    ]
    ws.append(headers)
    for code in sorted(by_code, key=lambda x: (decoded.get(x, RavenolResult(x)).manufacturer or "яяя", x)):
        items = by_code[code]
        result = decoded.get(code, RavenolResult(code=code, status="Не проверен"))
        pages = ", ".join(map(str, sorted({m.page for m in items})))
        ws.append([
            code, items[0].code_type, result.manufacturer or "Не определён", result.model,
            result.engine, result.year, result.status, result.source_method, result.result_url,
            len(items), pages, items[0].author, items[0].text[:1000],
        ])
    style_sheet(ws)
    ws.column_dimensions["M"].width = 75

    summary = wb.create_sheet("Сводка производителей")
    summary.append(["Автопроизводитель", "Уникальных VIN / Frame", "Упоминаний в комментариях"])
    manufacturer_codes: dict[str, set[str]] = defaultdict(set)
    manufacturer_mentions: Counter[str] = Counter()
    for code, items in by_code.items():
        manufacturer = decoded.get(code, RavenolResult(code)).manufacturer or "Не определён"
        manufacturer_codes[manufacturer].add(code)
        manufacturer_mentions[manufacturer] += len(items)
    for manufacturer in sorted(manufacturer_codes, key=lambda k: (-len(manufacturer_codes[k]), k)):
        summary.append([manufacturer, len(manufacturer_codes[manufacturer]), manufacturer_mentions[manufacturer]])
    style_sheet(summary)

    raw = wb.create_sheet("Все упоминания")
    raw.append(["VIN / Frame", "Тип", "Автопроизводитель", "Страница", "ID комментария", "Автор", "Текст комментария", "URL Drive2"])
    for m in sorted(mentions, key=lambda x: (x.page, x.comment_id, x.code)):
        manufacturer = decoded.get(m.code, RavenolResult(m.code)).manufacturer or "Не определён"
        raw.append([m.code, m.code_type, manufacturer, m.page, m.comment_id, m.author, m.text, m.source_url])
    style_sheet(raw)
    raw.column_dimensions["G"].width = 90
    raw.column_dimensions["H"].width = 55

    ravenol = wb.create_sheet("Проверка Ravenol")
    ravenol.append(["VIN / Frame", "Производитель", "Модель", "Двигатель", "Год", "Статус", "Метод", "URL", "Фрагмент ответа"])
    for code in sorted(decoded):
        r = decoded[code]
        ravenol.append([r.code, r.manufacturer, r.model, r.engine, r.year, r.status, r.source_method, r.result_url, r.response_excerpt])
    style_sheet(ravenol)
    ravenol.column_dimensions["I"].width = 90

    log_ws = wb.create_sheet("Журнал страниц")
    log_ws.append(["Страница", "URL", "Статус", "Комментариев найдено", "Примечание"])
    for item in logs:
        log_ws.append([item["page"], item["url"], item["status"], item["comments"], item["note"]])
    style_sheet(log_ws)

    method = wb.create_sheet("Методика")
    method.append(["Параметр", "Значение"])
    method.append(["Источник комментариев", base_url])
    method.append(["Каталог проверки", "https://podbor.ravenol.ru/"])
    method.append(["Дата и время формирования, UTC", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    method.append(["Страниц обработано", len(logs)])
    method.append(["Уникальных кодов", len(by_code)])
    method.append(["Всего упоминаний", len(mentions)])
    method.append(["Правило VIN", "17 символов A-H, J-N, P, R-Z и 0-9; I/O/Q исключены; пробелы и дефисы внутри нормализуются"])
    method.append(["Frame-коды", "Добавляются только при наличии в комментарии слов VIN/frame/вин/фрейм/кузов/шасси"])
    method.append(["Ограничение", "Автопроизводитель считается подтверждённым Ravenol, когда он прочитан из результата/URL. WMI используется только как явно отмеченный резерв после запроса Ravenol."])
    method.append(["Антибот", "CAPTCHA и ограничения доступа не обходятся; ошибки фиксируются в журнале"])
    style_sheet(method)
    method.column_dimensions["A"].width = 38
    method.column_dimensions["B"].width = 110

    wb.save(output)


async def main_async(args: argparse.Namespace) -> int:
    output = Path(args.output)
    debug_dir = output.parent / "debug"
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--disable-dev-shm-usage", "--no-sandbox"])
        context = await browser.new_context(
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            viewport={"width": 1366, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
            ),
            extra_http_headers={"Accept-Language": "ru-RU,ru;q=0.9,en;q=0.7"},
        )
        try:
            mentions, logs = await collect_drive2_mentions(context, args.drive2_url.rstrip("/") + "/", args.last_page)
            unique_codes = sorted({m.code for m in mentions})
            print(f"Collected {len(mentions)} mentions, {len(unique_codes)} unique codes", flush=True)
            decoded = await decode_all(context, unique_codes, debug_dir)
            write_excel(output, args.drive2_url, mentions, logs, decoded)
            print(f"Excel written: {output}", flush=True)
            return 0
        finally:
            await context.close()
            await browser.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Collect public VIN/Frame mentions from a Drive2 thread and verify them via Ravenol.")
    parser.add_argument("--drive2-url", default="https://www.drive2.ru/o/b/517173162262135362/")
    parser.add_argument("--last-page", type=int, default=144)
    parser.add_argument("--output", default="reports/drive2_ravenol_vin_summary.xlsx")
    args = parser.parse_args()
    return asyncio.run(main_async(args))


if __name__ == "__main__":
    raise SystemExit(main())
