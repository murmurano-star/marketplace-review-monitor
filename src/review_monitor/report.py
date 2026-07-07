from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analysis import summary, weekly_trend
from .models import Review

HEADERS = [
    "Площадка", "Дата", "Бренд", "Магазин продавца", "Товар", "ID товара",
    "Оценка", "Тональность", "Текст", "Достоинства", "Недостатки", "Боли",
    "Есть фото", "Есть видео", "Ссылка", "ID отзыва",
]


def _row(r: Review) -> list[Any]:
    return [
        r.platform, r.review_date.isoformat(), r.brand, r.seller, r.product_name,
        r.product_id, r.rating, r.sentiment, r.text, r.pros, r.cons,
        "; ".join(r.pains), "Да" if r.has_photo else "Нет",
        "Да" if r.has_video else "Нет", r.source_url, r.review_id,
    ]


def _style(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        values = [str(c.value or "") for c in ws[get_column_letter(col)]]
        ws.column_dimensions[get_column_letter(col)].width = min(max(max(map(len, values), default=10) + 2, 12), 45)
    if ws.max_row > 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def write_excel(reviews: list[Review], path: str | Path, date_from=None, date_to=None) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    s = summary(reviews)
    ws.append(["Показатель", "Значение"])
    if date_from and date_to:
        ws.append(["Период", f"{date_from.isoformat()} — {date_to.isoformat()}"])
    ws.append(["Всего отзывов", s["total"]])
    ws.append(["Средний рейтинг", s["average_rating"]])
    ws.append(["Негативных отзывов", s["negative_count"]])
    ws.append(["Доля негатива, %", s["negative_share"]])
    ws.append(["Отзывы с медиа", s["media_count"]])
    ws.append([])
    ws.append(["Боль", "Количество"])
    for label, count in s["pains"].most_common():
        ws.append([label, count])
    _style(ws)

    for name, items in {
        "Все отзывы": reviews,
        "Негатив": [r for r in reviews if r.rating <= 2],
        "Отзывы с медиа": [r for r in reviews if r.has_photo or r.has_video],
    }.items():
        tab = wb.create_sheet(name)
        tab.append(HEADERS)
        for review in items:
            tab.append(_row(review))
        _style(tab)

    pain_ws = wb.create_sheet("Боли")
    pain_ws.append(["Боль", "Количество"])
    for label, count in Counter(p for r in reviews for p in r.pains).most_common():
        pain_ws.append([label, count])
    _style(pain_ws)

    trend_ws = wb.create_sheet("Недельная динамика")
    trend_ws.append(["Неделя", "Бренд", "Магазин", "Отзывы", "Средний рейтинг", "Доля негатива, %", "С медиа"])
    for item in weekly_trend(reviews):
        trend_ws.append([item["week"], item["brand"], item["seller"], item["reviews"], item["average_rating"], item["negative_share"], item["with_media"]])
    _style(trend_ws)
    wb.save(path)


def write_markdown(reviews: list[Review], path: str | Path, date_from=None, date_to=None) -> None:
    s = summary(reviews)
    lines = ["# Мониторинг отзывов", ""]
    if date_from and date_to:
        lines += [f"Период: **{date_from.date()} — {date_to.date()}**", ""]
    lines += [
        f"- Отзывов: **{s['total']}**",
        f"- Средний рейтинг: **{s['average_rating']}**",
        f"- Доля негатива: **{s['negative_share']}%**",
        f"- Доля отзывов с медиа: **{s['media_share']}%**",
        "", "## Топ болей", "", "| Боль | Количество |", "|---|---:|",
    ]
    for label, count in s["pains"].most_common(10):
        lines.append(f"| {label} | {count} |")
    Path(path).write_text("\n".join(lines), encoding="utf-8")


def write_dashboard(reviews: list[Review], path: str | Path, max_rows: int = 5000) -> None:
    rows = []
    for r in reviews[:max_rows]:
        text = (r.full_text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        rows.append(
            f"<tr><td>{r.review_date.date()}</td><td>{r.platform}</td><td>{r.brand}</td>"
            f"<td>{r.seller}</td><td>{r.product_name}</td><td>{r.rating}</td>"
            f"<td>{text}</td><td>{'; '.join(r.pains)}</td></tr>"
        )
    page = (
        "<!doctype html><html lang='ru'><meta charset='utf-8'><title>Отзывы</title>"
        "<style>body{font-family:Arial;margin:24px}table{border-collapse:collapse;width:100%}"
        "th,td{border:1px solid #ddd;padding:7px;vertical-align:top}th{background:#111827;color:white}</style>"
        "<h1>Отзывы Wildberries и Ozon</h1><table><thead><tr><th>Дата</th><th>Площадка</th>"
        "<th>Бренд</th><th>Магазин</th><th>Товар</th><th>Оценка</th><th>Отзыв</th><th>Боли</th>"
        "</tr></thead><tbody>" + "".join(rows) + "</tbody></table></html>"
    )
    Path(path).write_text(page, encoding="utf-8")
