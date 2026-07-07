from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import Collector
from ..models import Review
from ..utils import list_from_value, parse_datetime


class ImportFileCollector(Collector):
    def __init__(self, config: dict[str, Any]):
        self.path = Path(config["path"])
        self.default_platform = str(config.get("platform", "")).strip()
        self.source_id = str(config.get("id", self.path.stem))

    def collect(self, date_from: datetime, date_to: datetime) -> list[Review]:
        if not self.path.exists():
            raise FileNotFoundError(f"Не найден файл импорта: {self.path}")
        output: list[Review] = []
        for index, row in enumerate(self._read_rows(), start=2):
            review = self._normalize(row, index)
            if date_from <= review.review_date <= date_to:
                output.append(review)
        return sorted(output, key=lambda r: r.review_date, reverse=True)

    def _read_rows(self) -> list[dict[str, Any]]:
        suffix = self.path.suffix.lower()
        if suffix == ".csv":
            with self.path.open("r", encoding="utf-8-sig", newline="") as fh:
                return list(csv.DictReader(fh))
        if suffix in {".xlsx", ".xlsm"}:
            wb = load_workbook(self.path, read_only=True, data_only=True)
            ws = wb.active
            iterator = ws.iter_rows(values_only=True)
            headers = [str(x or "").strip() for x in next(iterator)]
            return [dict(zip(headers, row)) for row in iterator]
        raise ValueError(f"Поддерживаются CSV и XLSX, получено: {suffix}")

    def _normalize(self, row: dict[str, Any], index: int) -> Review:
        platform = str(row.get("platform") or self.default_platform).strip().lower()
        return Review(
            platform=platform,
            review_id=str(row.get("review_id") or f"{self.path.stem}-{index}"),
            review_date=parse_datetime(row.get("review_date")),
            rating=int(float(row.get("rating") or 0)),
            brand=str(row.get("brand") or "").strip(),
            seller=str(row.get("seller") or "").strip(),
            product_name=str(row.get("product_name") or "").strip(),
            product_id=str(row.get("product_id") or "").strip(),
            sku=str(row.get("sku") or "").strip(),
            text=str(row.get("text") or "").strip(),
            pros=str(row.get("pros") or "").strip(),
            cons=str(row.get("cons") or "").strip(),
            photo_urls=list_from_value(row.get("photo_urls")),
            video_urls=list_from_value(row.get("video_urls")),
            source_url=str(row.get("source_url") or "").strip(),
            source_id=self.source_id,
            raw={str(k): v for k, v in row.items()},
        )
